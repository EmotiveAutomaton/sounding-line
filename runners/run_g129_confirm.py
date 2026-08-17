"""G129 CONFIRMATORY — event-level choice recovery on ArgRewrite, the preregistered battery.

Runs the card in prereg/g129.py (Amendment 1 in force). Build order:

    --build-manifest    constructs BOTH populations and every candidate set ONCE,
                        deterministically (seed 31), into results/g129/manifest.json --
                        so all arms consume identical events, candidates, and per-arm ask
                        seeds, and no arm's rng stream can couple to another's (the pilot's
                        shared-stream defect, closed by design)
    --arm X             X in {recovery, blind, shuffle, brief, source, unchanged}
                        (Ollama reader, checkpointed per event, resumes across passes)
                        or X = change_block (CPU: the 19-dim declared baseline,
                        author-grouped CV, probabilities restricted to each candidate set)
    --verdict           computes the preregistered bands, the reader-vs-block McNemar, and
                        the fabrication/miss rates; writes results/g129/verdict.json with
                        every statistic the verdict rests on

Populations. FULL: fine labels with >= 30 events on the full extracted set, truth-balanced
(analytic floor 1/k). MATCHED: the G130b/G130c CEM strata (fresh seed), common support,
labels >= 20 within support, then truth-balanced WITHIN support (the L126 amendment: the
matched floor returns to analytic 1/k). UNCHANGED: identical aligned pairs from the
annotation workbooks, plus a same-size sample of real changed events as the symmetric
control; every A7 candidate set carries the explicit "no revision was made" option.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / "runners"))

RESULTS = REPO / "results" / "g129"
EVENTS = REPO / "results" / "arg_baselines" / "events.json"
ANNOT = REPO / "corpora" / "public" / "argrewrite" / "annotations"
MANIFEST = RESULTS / "manifest.json"
OLLAMA = "http://127.0.0.1:11434/api/generate"
MODEL = "qwen3.5:9b"
SEED = 31
K = 4
NO_REV = "no revision was made"
BRIEF = ("The student was asked to write a short argumentative essay taking and defending a "
         "position on self-driving cars, and then to revise it across drafts after feedback.")

COMMON = None


def rare_rate(text: str) -> float:
    global COMMON
    if COMMON is None:
        from wordfreq import top_n_list                               # noqa: PLC0415
        COMMON = set(top_n_list("en", 5000))
    ws = [w.strip(".,;:!?\"'()").lower() for w in text.split()]
    ws = [w for w in ws if w]
    return sum(w not in COMMON for w in ws) / len(ws) if ws else 0.0


def covs(e) -> list[float]:
    o, n = e["old"].split(), e["new"].split()
    os_, ns_ = set(w.lower() for w in o), set(w.lower() for w in n)
    return [len(ns_ - os_), len(os_ - ns_), len(n) - len(o),
            rare_rate(e["new"]) - rare_rate(e["old"]),
            min(len(o), 60), rare_rate(e["old"])]


def ask(prompt: str, seed: int) -> str:
    req = urllib.request.Request(OLLAMA, data=json.dumps(
        {"model": MODEL, "prompt": prompt, "stream": False, "think": False,
         "options": {"temperature": 0.0, "seed": seed, "num_predict": 30}}).encode(),
        headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=300) as r:
        resp = json.loads(r.read()).get("response", "")
    return re.sub(r"<think>.*?</think>", "", resp, flags=re.DOTALL).strip()


def delta_text(e) -> str:
    o, n = e["old"].lower().split(), e["new"].lower().split()
    added = [w for w in n if w not in set(o)][:25]
    removed = [w for w in o if w not in set(n)][:25]
    return (f"ADDED: {' '.join(added) or '(nothing)'}\n"
            f"REMOVED: {' '.join(removed) or '(nothing)'}\n"
            f"BEFORE: {e['old'][:250]}\nAFTER: {e['new'][:250]}")


def extract_unchanged() -> list[dict]:
    """Identical aligned pairs from the annotation workbooks: the A7 substrate."""
    import warnings                                                   # noqa: PLC0415
    warnings.filterwarnings("ignore")
    from openpyxl import load_workbook                                # noqa: PLC0415
    out = []
    for wb_path in sorted(ANNOT.rglob("*.xlsx")):
        am = re.search(r"argrewrite_(\d+)", wb_path.stem)
        author = am.group(1) if am else wb_path.stem
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
        except Exception:                                             # noqa: BLE001
            continue
        for ws in wb.worksheets:
            if "new" not in ws.title.lower():
                continue
            ws.reset_dimensions()
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

            def col(name):
                return next((i for i, h in enumerate(header) if name in h), None)

            it, ii, ip = col("sentence content"), col("identical"), col("purpose level 1")
            if it is None or ii is None:
                continue
            for r in rows[1:]:
                txt = str(r[it] or "").strip() if it < len(r) else ""
                raw = r[ii] if ii < len(r) else None
                try:
                    ident = float(raw) == 1.0        # sheets store 1.0/0.0, not yes/no
                except (TypeError, ValueError):
                    ident = str(raw).strip().lower() in ("yes", "true", "identical")
                purp = str(r[ip] or "").strip().lower() if ip is not None and ip < len(r) else ""
                if txt and ident and purp in ("", "none", "nan"):
                    out.append({"author": author, "old": txt[:400], "new": txt[:400]})
        wb.close()
    return out


def build_manifest() -> None:
    import numpy as np                                                # noqa: PLC0415

    rng = np.random.default_rng(SEED)
    events = json.loads(EVENTS.read_text(encoding="utf-8"))["events"]

    def balance(pool, min_count):
        labs = [e["fine"] for e in pool]
        kept = sorted({l for l in set(labs) if labs.count(l) >= min_count})
        pool = [e for e in pool if e["fine"] in kept]
        per = min(sum(1 for e in pool if e["fine"] == l) for l in kept) if kept else 0
        by: dict = {}
        for e in pool:
            by.setdefault(e["fine"], []).append(e)
        bal = [e for l in kept for e in
               [by[l][j] for j in rng.permutation(len(by[l]))[:per]]]
        return bal, kept

    # FULL population
    full, full_labs = balance(events, 30)

    # MATCHED population: same strata construction as G130b/G130c, fresh permutation seed
    all_sc = [e for e in events if e["coarse"] in ("surface", "content")]
    C = np.array([covs(e) for e in all_sc], float)
    yc = np.array([e["coarse"] == "content" for e in all_sc])
    Z = (C - C.mean(0)) / (C.std(0) + 1e-9)
    bins = np.stack([np.digitize(Z[:, j], np.quantile(Z[:, j], [1 / 3, 2 / 3]))
                     for j in range(Z.shape[1])], axis=1)
    strata: dict = {}
    for i, b in enumerate(bins):
        strata.setdefault(tuple(b), {"c": [], "s": []})["c" if yc[i] else "s"].append(i)
    keep = set()
    for st in strata.values():
        n = min(len(st["c"]), len(st["s"]))
        if n == 0:
            continue
        keep.update(rng.permutation(st["c"])[:n].tolist())
        keep.update(rng.permutation(st["s"])[:n].tolist())
    matched_support = [all_sc[i] for i in sorted(keep)]
    matched, matched_labs = balance(matched_support, 20)

    # UNCHANGED population + symmetric changed control (same size)
    unchanged = extract_unchanged()
    n_u = min(200, len(unchanged))
    unchanged = [unchanged[j] for j in rng.permutation(len(unchanged))[:n_u]]
    sym = [full[j] for j in rng.permutation(len(full))[:n_u]]

    def cand_sets(pop, labs, extra=None):
        out = []
        for e in pop:
            truth = e.get("fine")
            decoys = [l for l in labs if l != truth]
            picks = list(rng.choice(decoys, size=min(K - 1, len(decoys)), replace=False))
            cands = picks + ([truth] if truth else
                             list(rng.choice(decoys, size=1)))
            if extra:
                cands = cands + [extra]
            cands = [cands[j] for j in rng.permutation(len(cands))]
            out.append(cands)
        return out

    man = {
        "seed": SEED, "k": K, "model": MODEL, "prereg": "prereg/g129.py (Amendment 1)",
        "full": {"events": full, "labels": full_labs, "cands": cand_sets(full, full_labs),
                 "shuffle_truth": list(rng.permutation([e["fine"] for e in full]))},
        "matched": {"events": matched, "labels": matched_labs,
                    "cands": cand_sets(matched, matched_labs),
                    "n_required": 283, "n_actual": len(matched)},
        "unchanged": {"events": unchanged,
                      "cands": cand_sets(unchanged, full_labs, extra=NO_REV),
                      "sym_events": sym,
                      "sym_cands": cand_sets(sym, full_labs, extra=NO_REV)},
        "ask_seed_base": {"recovery": 1000, "blind": 2000, "shuffle": 3000,
                          "brief": 4000, "source": 5000, "unchanged": 6000,
                          "recovery_matched": 7000, "blind_matched": 8000},
    }
    RESULTS.mkdir(parents=True, exist_ok=True)
    MANIFEST.write_text(json.dumps(man), encoding="utf-8", newline="\n")
    print(f"manifest: full {len(full)} ({len(full_labs)} labels), matched {len(matched)} "
          f"(required 283), unchanged {len(unchanged)}+{len(sym)} sym")
    if len(matched) < 283:
        print("SHORTFALL: matched population under the powered n; the card's downgrade "
              "clause applies at verdict time")
    print(f"wrote {MANIFEST.relative_to(REPO)}")


def run_reader_arm(arm: str) -> None:
    import numpy as np                                                # noqa: PLC0415
    from soundingline.gpulock import acquire_gpu_lock                 # noqa: PLC0415

    acquire_gpu_lock(f"g129_{arm}")                    # sustained ollama serializes on the card
    man = json.loads(MANIFEST.read_text(encoding="utf-8"))
    rng = np.random.default_rng(SEED + 1000)
    base = man["ask_seed_base"][arm]
    pop = ("matched" if arm.endswith("_matched") else
           "unchanged" if arm == "unchanged" else "full")
    core = arm.replace("_matched", "")

    if arm == "unchanged":
        events = man["unchanged"]["events"] + man["unchanged"]["sym_events"]
        cands_all = man["unchanged"]["cands"] + man["unchanged"]["sym_cands"]
        kinds = ["unchanged"] * len(man["unchanged"]["events"]) + \
                ["sym"] * len(man["unchanged"]["sym_events"])
    else:
        events = man[pop]["events"]
        cands_all = man[pop]["cands"]
        kinds = ["event"] * len(events)

    part = RESULTS / f"{arm}_partial.jsonl"
    done = set()
    if part.exists():
        for line in part.read_text(encoding="utf-8").splitlines():
            done.add(json.loads(line)["i"])
    with part.open("a", encoding="utf-8", newline="\n") as fh:
        for i, (e, cands, kind) in enumerate(zip(events, cands_all, kinds)):
            if i in done:
                continue
            if core == "blind":
                body = ""
            elif core == "brief":
                body = f"THE ASSIGNMENT: {BRIEF}\n"
            elif core == "source":
                body = f"ORIGINAL SENTENCE: {e['old'][:250]}\n"
            else:                                    # recovery, shuffle, unchanged
                body = delta_text(e) + "\n"
            opts = "\n".join(f"- {l}" for l in cands)
            ans = ask("A student revised one sentence of an argumentative essay.\n" + body +
                      f"Which revision purpose fits best? Answer with exactly one label "
                      f"from:\n{opts}\nLabel:", seed=base + i).lower()
            got = [l for l in cands if l in ans]
            pick = got[0] if len(got) == 1 else str(rng.choice(got or cands))
            truth = (man["full"]["shuffle_truth"][i] if core == "shuffle"
                     else e.get("fine", NO_REV if kind == "unchanged" else None))
            fh.write(json.dumps({"i": i, "kind": kind, "truth": truth, "pick": pick}) + "\n")
            fh.flush()

    rows = [json.loads(x) for x in part.read_text(encoding="utf-8").splitlines()]
    if arm == "unchanged":
        u = [r for r in rows if r["kind"] == "unchanged"]
        s = [r for r in rows if r["kind"] == "sym"]
        out = {"arm": arm, "n_unchanged": len(u), "n_sym": len(s),
               "fabrication_rate": round(sum(r["pick"] != NO_REV for r in u) / max(len(u), 1), 4),
               "sym_miss_rate": round(sum(r["pick"] == NO_REV for r in s) / max(len(s), 1), 4),
               "sym_accuracy": round(sum(r["pick"] == r["truth"] for r in s) / max(len(s), 1), 4)}
    else:
        acc = sum(r["truth"] == r["pick"] for r in rows) / max(len(rows), 1)
        k_eff = K
        out = {"arm": arm, "n": len(rows), "k": k_eff, "accuracy": round(acc, 4),
               "analytic_floor": round(1 / k_eff, 4)}
    (RESULTS / f"{arm}.json").write_text(json.dumps(out, indent=1),
                                         encoding="utf-8", newline="\n")
    print(json.dumps(out))


def run_change_block() -> None:
    import numpy as np                                                # noqa: PLC0415
    from run_arg_replication import change_features                   # noqa: PLC0415
    from sklearn.ensemble import HistGradientBoostingClassifier       # noqa: PLC0415

    man = json.loads(MANIFEST.read_text(encoding="utf-8"))
    out_all = {}
    for pop in ("full", "matched"):
        events, cands_all = man[pop]["events"], man[pop]["cands"]
        labs = man[pop]["labels"]
        X = np.array([change_features(e["old"], e["new"]) for e in events], float)
        y = np.array([labs.index(e["fine"]) for e in events])
        authors = np.array([e["author"] for e in events])
        picks = [None] * len(events)
        for held in sorted(set(authors)):
            tr, te = authors != held, authors == held
            if te.sum() == 0 or len(set(y[tr])) < 2:
                continue
            clf = HistGradientBoostingClassifier(max_iter=300, random_state=SEED)
            clf.fit(X[tr], y[tr])
            proba = clf.predict_proba(X[te])
            classes = list(clf.classes_)
            for row_p, gi in zip(proba, np.where(te)[0]):
                cand_idx = [labs.index(c) for c in cands_all[gi]]
                scores = [row_p[classes.index(ci)] if ci in classes else 0.0
                          for ci in cand_idx]
                picks[gi] = cands_all[gi][int(np.argmax(scores))]
        rows = [{"i": i, "truth": e["fine"], "pick": p}
                for i, (e, p) in enumerate(zip(events, picks)) if p is not None]
        acc = sum(r["truth"] == r["pick"] for r in rows) / max(len(rows), 1)
        out_all[pop] = {"n": len(rows), "accuracy": round(acc, 4)}
        (RESULTS / f"change_block_{pop}_partial.jsonl").write_text(
            "\n".join(json.dumps(r) for r in rows) + "\n", encoding="utf-8", newline="\n")
    out = {"arm": "change_block", **out_all}
    (RESULTS / "change_block.json").write_text(json.dumps(out, indent=1),
                                               encoding="utf-8", newline="\n")
    print(json.dumps(out, indent=1))


def verdict() -> None:
    from scipy.stats import binomtest                                 # noqa: PLC0415

    man = json.loads(MANIFEST.read_text(encoding="utf-8"))
    get = lambda a: json.loads((RESULTS / f"{a}.json").read_text(encoding="utf-8"))  # noqa: E731
    rows_of = lambda a: [json.loads(x) for x in                        # noqa: E731
                         (RESULTS / f"{a}_partial.jsonl").read_text(encoding="utf-8").splitlines()]

    out = {"prereg": "prereg/g129.py (Amendment 1)", "gates": {}, "verdicts": {}}

    # VOID gates: blind within CI of 1/k, shuffle at chance
    for gate_arm in ("blind", "shuffle", "blind_matched"):
        g = get(gate_arm)
        bt = binomtest(round(g["accuracy"] * g["n"]), g["n"], 1 / g["k"])
        out["gates"][gate_arm] = {"accuracy": g["accuracy"], "n": g["n"],
                                  "p_vs_floor": round(bt.pvalue, 5),
                                  "VOID_if": "p < 0.05", "void": bool(bt.pvalue < 0.05)}

    # H-A full-set margin over analytic floor
    rec = get("recovery")
    m_full = rec["accuracy"] - 1 / rec["k"]
    out["verdicts"]["H-A_full"] = {
        "margin": round(m_full, 4),
        "band": "REPLICATES" if m_full >= 0.15 else "PARTIAL" if m_full >= 0.08 else "FAILS"}

    # H-B matched margin over analytic floor (balanced within support, L126 amendment)
    recm = get("recovery_matched")
    m_m = recm["accuracy"] - 1 / recm["k"]
    out["verdicts"]["H-B_matched"] = {
        "margin": round(m_m, 4), "n": recm["n"], "n_required": man["matched"]["n_required"],
        "powered": recm["n"] >= man["matched"]["n_required"],
        "band": "SURVIVES" if m_m >= 0.08 else "WEAKENED" if m_m >= 0.04 else "COLLAPSED"}

    # H-C reader vs change block, exact McNemar on identical full-population events
    rr = {r["i"]: r for r in rows_of("recovery")}
    br = {r["i"]: r for r in rows_of("change_block_full")}
    both = sorted(set(rr) & set(br))
    b01 = sum(rr[i]["truth"] == rr[i]["pick"] and br[i]["truth"] != br[i]["pick"] for i in both)
    b10 = sum(rr[i]["truth"] != rr[i]["pick"] and br[i]["truth"] == br[i]["pick"] for i in both)
    mc = binomtest(b01, b01 + b10, 0.5) if (b01 + b10) else None
    call = ("TIES" if mc is None or mc.pvalue >= 0.05 else
            "BEATS" if b01 > b10 else "LOSES")
    out["verdicts"]["H-C_reader_vs_block"] = {
        "reader_only_correct": b01, "block_only_correct": b10, "n_both": len(both),
        "mcnemar_p": round(mc.pvalue, 5) if mc else None, "call": call}

    # A7 fabrication
    u = get("unchanged")
    out["verdicts"]["A7"] = {**u, "band": "CLEAN" if u["fabrication_rate"] <= 0.10
                             else "OVER-READS"}

    # context arms, reported
    for a in ("brief", "source"):
        out[a] = get(a)

    dest = RESULTS / "verdict.json"
    dest.write_text(json.dumps(out, indent=1), encoding="utf-8", newline="\n")
    print(json.dumps(out, indent=1))
    print(f"wrote {dest.relative_to(REPO)}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build-manifest", action="store_true")
    ap.add_argument("--arm", choices=["recovery", "blind", "shuffle", "brief", "source",
                                      "unchanged", "recovery_matched", "blind_matched",
                                      "change_block"])
    ap.add_argument("--verdict", action="store_true")
    args = ap.parse_args()
    if args.build_manifest:
        build_manifest()
    elif args.arm == "change_block":
        run_change_block()
    elif args.arm:
        run_reader_arm(args.arm)
    elif args.verdict:
        verdict()
    else:
        ap.error("pick a mode")


if __name__ == "__main__":
    main()
