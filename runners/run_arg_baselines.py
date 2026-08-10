"""G136 — recreate ArgRewrite's own published task before pushing past it (Phase 1).

The corpus paper reports revision-purpose classification at two grains (coarse surface/content;
fine purposes). This reproduces that task with our tooling so the published numbers become our
known answer, and its event extraction doubles as the choice-event dataset G129 runs on.

Arms (one stage each in the queue):
    --arm extract    parse every annotation workbook into per-revision events
                     (old sentence, new sentence, fine purpose, coarse class, author, cycle)
    --arm features   bag-of-words + diff features, logistic classifier, AUTHOR-split CV,
                     coarse and fine, confusion matrices
    --arm reader --cycle 12|23 --grain coarse|fine
                     the local reader model classifies each revision from its old->new pair
                     against the label set; checkpointed per revision, so the 120-minute stage
                     timeout only pauses it and the next queue pass resumes

Recreation gate: our numbers stand next to the paper's reported baselines (comparison recorded
in FINDINGS when the paper's tables are fetched and read; nothing is assumed from memory).
Author-split is enforced everywhere; hundreds of revisions by one writer are not independent.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))

RESULTS = REPO / "results" / "arg_baselines"
EVENTS = RESULTS / "events.json"
ANNOT = REPO / "corpora" / "public" / "argrewrite" / "annotations"
OLLAMA = "http://127.0.0.1:11434/api/generate"
MODEL = "qwen3.5:9b"

SURFACE = ("conventions", "grammar", "spelling", "word-usage", "word usage", "clarity",
           "organization", "fluency")
CONTENT = ("claim", "evidence", "reasoning", "rebuttal", "warrant", "counter", "idea",
           "content development", "precision")


def ask(prompt: str, seed: int) -> str:
    req = urllib.request.Request(OLLAMA, data=json.dumps(
        {"model": MODEL, "prompt": prompt, "stream": False, "think": False,
         "options": {"temperature": 0.0, "seed": seed, "num_predict": 40}}).encode(),
        headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=300) as r:
        resp = json.loads(r.read()).get("response", "")
    return re.sub(r"<think>.*?</think>", "", resp, flags=re.DOTALL).strip()


def extract() -> list[dict]:
    # real schema: two sheets, "Old Draft" and "New Draft", each row a sentence with
    # Sentence Index / Sentence Content / Aligned Index / Identical? / Revision Purpose Level 0
    # (Level 1 further right where present). New-sheet rows aligned to old give (old, new) pairs;
    # unaligned rows are additions; annotated old-sheet rows with no alignment are deletions.
    import warnings                                                   # noqa: PLC0415
    warnings.filterwarnings("ignore")
    from openpyxl import load_workbook                                # noqa: PLC0415

    def sheet_table(ws):
        ws.reset_dimensions()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(name):
            return next((i for i, h in enumerate(header) if name in h), None)

        idx = {"i": col("sentence index"), "txt": col("sentence content"),
               "al": col("aligned index"), "ident": col("identical"),
               "p1": col("purpose level 1"), "p0": col("purpose level 0")}
        if idx["i"] is None or idx["txt"] is None:
            return None
        out = []
        for r in rows[1:]:
            def cell(i):
                return r[i] if i is not None and i < len(r) else None
            out.append({k: cell(v) for k, v in idx.items()})
        return out

    events = []
    for wb_path in sorted(ANNOT.rglob("*.xlsx")):
        m = re.search(r"[\\/](12|23)[\\/]", str(wb_path))
        cyc = m.group(1) if m else "??"
        am = re.search(r"argrewrite_(\d+)", wb_path.stem)
        author = am.group(1) if am else wb_path.stem
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
        except Exception:
            continue
        sheets = {ws.title.lower(): sheet_table(ws) for ws in wb.worksheets}
        wb.close()
        old = next((v for k, v in sheets.items() if "old" in k and v), None)
        new = next((v for k, v in sheets.items() if "new" in k and v), None)
        if not new:
            continue
        old_by_i = {}
        for r in (old or []):
            try:
                old_by_i[int(float(r["i"]))] = str(r["txt"] or "").strip()
            except (TypeError, ValueError):
                continue

        def emit(fine, old_s, new_s):
            fine = str(fine).strip().lower()
            if not fine or fine in ("none", "nan"):
                return
            coarse = ("surface" if any(t in fine for t in SURFACE) else
                      "content" if any(t in fine for t in CONTENT) else "other")
            events.append({"author": author, "cycle": cyc, "fine": fine, "coarse": coarse,
                           "old": (old_s or "")[:400], "new": (new_s or "")[:400]})

        for r in new:
            fine = r["p1"] or r["p0"]
            if not fine:
                continue
            al = None
            try:
                al = int(float(str(r["al"]).split(",")[0]))
            except (TypeError, ValueError):
                pass
            emit(fine, old_by_i.get(al, ""), str(r["txt"] or "").strip())
        for r in (old or []):
            fine = r["p1"] or r["p0"]
            has_align = str(r["al"] or "").strip() not in ("", "None")
            if fine and not has_align:
                emit(fine, str(r["txt"] or "").strip(), "")
    return events


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--arm", required=True,
                    choices=["extract", "features", "reader"])
    ap.add_argument("--cycle", default="12")
    ap.add_argument("--grain", default="coarse", choices=["coarse", "fine"])
    args = ap.parse_args()
    RESULTS.mkdir(parents=True, exist_ok=True)

    if args.arm == "extract":
        events = extract()
        from collections import Counter                               # noqa: PLC0415
        fines = Counter(e["fine"] for e in events)
        print(f"{len(events)} events, {len({e['author'] for e in events})} authors, "
              f"{len(fines)} fine labels; top: {fines.most_common(8)}")
        if len(events) < 500:
            print(">>> TOO FEW EVENTS -- schema drift; no cache written")
            sys.exit(1)
        EVENTS.write_text(json.dumps({"events": events}, indent=1),
                          encoding="utf-8", newline="\n")
        print(f"wrote {EVENTS.relative_to(REPO)}")
        return

    events = json.loads(EVENTS.read_text(encoding="utf-8"))["events"]

    if args.arm == "features":
        import numpy as np                                            # noqa: PLC0415
        from sklearn.feature_extraction.text import TfidfVectorizer   # noqa: PLC0415
        from sklearn.linear_model import LogisticRegression           # noqa: PLC0415
        from sklearn.model_selection import GroupKFold                # noqa: PLC0415
        from sklearn.metrics import confusion_matrix, f1_score        # noqa: PLC0415

        out = {}
        for grain in ("coarse", "fine"):
            labs = [e[grain] for e in events]
            keep_lab = {l for l in set(labs) if labs.count(l) >= 30}
            idx = [i for i, l in enumerate(labs) if l in keep_lab]

            def diff_text(e):
                o, n = set(e["old"].lower().split()), set(e["new"].lower().split())
                added = " ".join(f"ADD_{w}" for w in sorted(n - o))
                removed = " ".join(f"DEL_{w}" for w in sorted(o - n))
                return f"{added} {removed} || {e['old']} || {e['new']}"

            X_txt = [diff_text(events[i]) for i in idx]
            y = [labs[i] for i in idx]
            groups = [events[i]["author"] for i in idx]
            vec = TfidfVectorizer(max_features=5000, ngram_range=(1, 2))
            X = vec.fit_transform(X_txt)
            preds = [None] * len(y)
            for tr, te in GroupKFold(n_splits=5).split(X, y, groups):
                clf = LogisticRegression(max_iter=2000).fit(X[tr], [y[i] for i in tr])
                for i, p in zip(te, clf.predict(X[te])):
                    preds[i] = p
            f1 = f1_score(y, preds, average="macro")
            labs_sorted = sorted(keep_lab)
            cm = confusion_matrix(y, preds, labels=labs_sorted).tolist()
            out[grain] = {"n": len(y), "n_labels": len(keep_lab), "macro_f1": float(f1),
                          "labels": labs_sorted, "confusion": cm,
                          "chance_f1_lowerbound": 1.0 / len(keep_lab)}
            print(f"features {grain}: n={len(y)} labels={len(keep_lab)} macro-F1={f1:.3f}")
        (RESULTS / "features.json").write_text(json.dumps(out, indent=1),
                                               encoding="utf-8", newline="\n")
        print(f"wrote {(RESULTS / 'features.json').relative_to(REPO)}")
        return

    # reader arm, checkpointed
    sub = [e for e in events if e["cycle"] == args.cycle]
    labs = sorted({e[args.grain] for e in sub
                   if args.grain == "coarse" or
                   sum(1 for x in sub if x[args.grain] == e[args.grain]) >= 30})
    part = RESULTS / f"reader_{args.cycle}_{args.grain}_partial.jsonl"
    done = set()
    if part.exists():
        for line in part.read_text(encoding="utf-8").splitlines():
            done.add(json.loads(line)["i"])
    print(f"cycle {args.cycle} {args.grain}: {len(sub)} events, {len(labs)} labels, "
          f"{len(done)} already done")
    with part.open("a", encoding="utf-8", newline="\n") as fh:
        for i, e in enumerate(sub):
            if i in done or e[args.grain] not in labs:
                continue
            opts = "\n".join(f"- {l}" for l in labs)
            ans = ask(
                "A student revised one sentence of an essay.\n"
                f"BEFORE: {e['old']}\nAFTER: {e['new']}\n"
                f"Which revision purpose fits best? Answer with exactly one label from:\n{opts}\n"
                "Label:", seed=100 + i)
            ans_l = ans.lower()
            pick = next((l for l in labs if l in ans_l), "NONE")
            fh.write(json.dumps({"i": i, "truth": e[args.grain], "pick": pick,
                                 "author": e["author"]}) + "\n")
            fh.flush()
            if (i + 1) % 50 == 0:
                print(f"  {i + 1}/{len(sub)}", flush=True)

    rows = [json.loads(x) for x in part.read_text(encoding="utf-8").splitlines()]
    hits = sum(r["truth"] == r["pick"] for r in rows)
    acc = hits / max(len(rows), 1)
    per_author: dict[str, list[int]] = {}
    for r in rows:
        per_author.setdefault(r["author"], []).append(r["truth"] == r["pick"])
    out = {"cycle": args.cycle, "grain": args.grain, "n": len(rows),
           "n_labels": len(labs), "accuracy": acc, "chance": 1.0 / max(len(labs), 1),
           "per_author_acc": {a: sum(v) / len(v) for a, v in per_author.items()}}
    (RESULTS / f"reader_{args.cycle}_{args.grain}.json").write_text(
        json.dumps(out, indent=1), encoding="utf-8", newline="\n")
    print(f"reader {args.cycle}/{args.grain}: acc {acc:.3f} vs chance {out['chance']:.3f} "
          f"on {len(rows)}")
    print(f"wrote {(RESULTS / f'reader_{args.cycle}_{args.grain}.json').relative_to(REPO)}")


if __name__ == "__main__":
    main()
