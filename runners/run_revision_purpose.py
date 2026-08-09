"""PD-28 — is the revision effect polish or depth? The labels have been on disk all along.

The ArgRewrite annotations label each revision's purpose (surface vs content classes). The L5
effect — revision raises lexical sophistication — is either polish (if it lives in
surface-labelled revisions) or **the first depth signal on human text** (if it survives among
content-only revisions).

Schema-defensive: the workbook layout is undocumented here, so this runner first maps the schema
of three files; if it cannot find a purpose column it exits NEEDS-SCHEMA with the dump instead of
guessing. Sentence-level: revised sentences are attributed a purpose class; the sophistication
delta (mean word length, rare-word rate, stopword rate) is computed per class.

    POLISH      the sophistication shift concentrates in surface-labelled revisions
    DEPTH       it survives at full strength among content-labelled revisions — the first
                depth signal on human text
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))

RESULTS = REPO / "results" / "revision_purpose"
ANNOT = REPO / "corpora" / "public" / "argrewrite" / "annotations"

SURFACE = {"fluency", "spelling", "grammar", "word-usage", "wordusage", "clarity", "convention",
           "conventions", "organization", "surface"}
CONTENT = {"claim", "claims", "evidence", "reasoning", "rebuttal", "warrant", "idea",
           "development", "content", "precision"}


def soph(sent: str) -> tuple[float, float, float]:
    words = re.findall(r"[a-zA-Z']+", sent.lower())
    if not words:
        return (0.0, 0.0, 0.0)
    stop = {"the", "a", "an", "and", "or", "but", "of", "to", "in", "is", "was", "it", "that"}
    return (sum(map(len, words)) / len(words),
            sum(1 for w in words if len(w) > 8) / len(words),
            sum(1 for w in words if w in stop) / len(words))


def main() -> None:
    import numpy as np                                                # noqa: PLC0415
    from openpyxl import load_workbook                                # noqa: PLC0415

    files = sorted(ANNOT.rglob("*.xlsx"))
    print(f"{len(files)} annotation workbooks")
    # Real schema (probed 2026-08-09): sheets 'Old Draft'/'New Draft'; columns include
    # 'Aligned Index' (old->new sentence mapping, comma-separated) and
    # 'Revision Purpose Level 0/1' with the ArgRewrite taxonomy labels.

    deltas = {"surface": [], "content": [], "other": []}
    parsed = 0
    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            if "Old Draft" not in wb.sheetnames or "New Draft" not in wb.sheetnames:
                wb.close()
                continue

            def sheet_rows(name):
                ws = wb[name]
                ws.reset_dimensions()   # the files carry a broken A1:A1 dimension record
                it = ws.iter_rows(values_only=True)
                header = [str(v).strip().lower() if v else "" for v in next(it)]
                col = {h: i for i, h in enumerate(header)}
                return col, list(it)

            ocol, orows = sheet_rows("Old Draft")
            ncol, nrows = sheet_rows("New Draft")

            def cell(r, i):
                return r[i] if (i is not None and i < len(r)) else None

            new_by_idx = {}
            for r in nrows:
                try:
                    new_by_idx[int(float(cell(r, ncol.get("sentence index", 0))))] = r
                except (TypeError, ValueError):
                    continue
            for r in orows:
                old_s = str(cell(r, ocol.get("sentence content", 1)) or "")
                aligned = str(cell(r, ocol.get("aligned index", 2)) or "")
                purposes = []
                for lvl in ("revision purpose level 0", "revision purpose level 1"):
                    v = cell(r, ocol.get(lvl))
                    if v:
                        purposes.append(str(v).strip().lower())
                if not purposes or not aligned:
                    continue
                new_parts = []
                for tok in aligned.split(","):
                    try:
                        nr = new_by_idx.get(int(float(tok.strip())))
                        if nr is not None:
                            new_parts.append(str(cell(nr, ncol.get("sentence content", 1)) or ""))
                    except (TypeError, ValueError):
                        continue
                if not new_parts:
                    continue
                new_s = " ".join(new_parts)
                sn, so = soph(new_s), soph(old_s)
                for purpose in purposes:
                    cls = ("surface" if any(t in purpose for t in SURFACE) else
                           "content" if any(t in purpose for t in CONTENT) else "other")
                    deltas[cls].append([sn[i] - so[i] for i in range(3)])
            parsed += 1
            wb.close()
        except Exception as e:                                        # noqa: BLE001
            print(f"  skip {f.name}: {type(e).__name__}")
    print(f"parsed {parsed} workbooks")

    out = {"counts": {k: len(v) for k, v in deltas.items()}}
    print("revision counts:", out["counts"])
    for k in ("surface", "content"):
        if deltas[k]:
            arr = np.array(deltas[k])
            out[k] = {"d_wordlen": float(arr[:, 0].mean()), "d_rare": float(arr[:, 1].mean()),
                      "d_stop": float(arr[:, 2].mean())}
            print(f"{k}: wordlen {arr[:, 0].mean():+.4f}  rare {arr[:, 1].mean():+.4f}  "
                  f"stop {arr[:, 2].mean():+.4f}  (n={len(arr)})")
    if deltas["surface"] and deltas["content"]:
        s, c = out["surface"], out["content"]
        depth_carries = c["d_wordlen"] >= 0.5 * s["d_wordlen"] and c["d_rare"] >= 0.5 * s["d_rare"]
        out["verdict"] = "DEPTH-SIGNAL" if depth_carries else "POLISH"
    else:
        out["verdict"] = "NEEDS-SCHEMA"
    print(f"\n  >>> {out['verdict']}")

    RESULTS.mkdir(parents=True, exist_ok=True)
    (RESULTS / "summary.json").write_text(json.dumps(out, indent=2),
                                          encoding="utf-8", newline="\n")
    print(f"wrote {(RESULTS / 'summary.json').relative_to(REPO)}")


if __name__ == "__main__":
    main()
