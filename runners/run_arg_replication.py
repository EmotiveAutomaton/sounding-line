"""G136-exact — recreate the ArgRewrite paper's classification numbers, not its conclusions.

THE CAPABILITY PASS. The published sentential baselines (Table 6/7, fetched and read 2026-08-10):

    binary   Majority .37/.58   Features .90/.90   USE .92/.92   Features+USE .93/.93
    fine     Majority .05/.29   Features .44/.58   USE .49/.62   Features+USE .51/.63
                                                   (avg unweighted F1 / accuracy, 5-fold CV)

Their exact recipe, reproduced: features = sentence length, sentence position, POS-tag term
frequencies (spaCy), transition-word term frequencies; embeddings = Universal Sentence Encoder
(transformer variant) on the <old, new> pair; classifier = XGBoost, grid n_estimators in
{250,500,750,1000} x max_depth in {3,4,5} x lr in {.1,.05,.01}; 5-fold CV, average unweighted
F-score and accuracy. No author grouping is mentioned in the paper, so the folds here are plain
shuffled KFold, matching their protocol rather than improving on it.

PASS = the Features+USE cells match the published numbers at two decimals. Anything further off
implies our modeling of their pipeline is wrong somewhere, and the deltas table is the search map.

Known underdeterminations, resolved to the most standard reading and recorded: the USE pair
combination (concatenation of the two 512-d vectors), POS/transition features computed on old and
new and concatenated, multi-purpose cells split into one example per purpose, grid selected on
the same 5-fold mean (the common practice when no nested tuning is described).
"""

from __future__ import annotations

import json
import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))

RESULTS = REPO / "results" / "arg_baselines"
ANNOT = REPO / "corpora" / "public" / "argrewrite" / "annotations"

TARGETS = {
    "binary": {"majority": (0.37, 0.58), "features": (0.90, 0.90),
               "use": (0.92, 0.92), "features_use": (0.93, 0.93)},
    "fine": {"majority": (0.05, 0.29), "features": (0.44, 0.58),
             "use": (0.49, 0.62), "features_use": (0.51, 0.63)},
}
FINE9 = {
    "word-usage/clarity": "word_usage", "conventions/grammar/spelling": "grammar_spelling",
    "organization": "organization", "claims/ideas": "claim",
    "warrant/reasoning/backing": "reasoning", "evidence": "evidence",
    "rebuttal/reservation": "rebuttal", "precision": "precision",
    "general content development": "general_content",
}
SURFACE9 = {"word_usage", "grammar_spelling", "organization"}
TRANSITIONS = ("however", "therefore", "moreover", "furthermore", "consequently", "although",
               "because", "since", "thus", "hence", "meanwhile", "nevertheless", "instead",
               "additionally", "also", "finally", "first", "second", "third", "then", "next",
               "for example", "for instance", "in addition", "in conclusion", "in contrast",
               "on the other hand", "as a result", "in fact", "indeed", "overall", "similarly",
               "specifically", "accordingly", "besides", "still", "yet", "so", "but", "and")


TAGSET19 = ("ADJ", "ADP", "ADV", "AUX", "CCONJ", "CONJ", "DET", "INTJ", "NOUN", "NUM",
            "PART", "PRON", "PROPN", "PUNCT", "SCONJ", "SYM", "VERB", "X", "SPACE")


def change_features(old: str, new: str) -> list[float]:
    """Nineteen surface measures of WHAT CHANGED between the two sentences.

    The pair tasks' load-bearing signal (L85): a revision is defined by its delta, and
    neither a bare concatenation of sentence embeddings nor per-sentence counts state the
    delta anywhere. Half these pairs have one empty side, so 'how much changed' is most of
    the recoverable structure.
    """
    from difflib import SequenceMatcher                               # noqa: PLC0415
    ot, nt = old.split(), new.split()
    so, sn = set(ot), set(nt)
    union = len(so | sn) or 1
    inter = len(so & sn)
    char_ratio = SequenceMatcher(None, old, new).ratio() if (old or new) else 1.0
    sm = SequenceMatcher(None, ot, nt)
    tok_ratio = sm.ratio() if (ot or nt) else 1.0
    ins = dele = rep = eq = 0
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "insert":
            ins += j2 - j1
        elif tag == "delete":
            dele += i2 - i1
        elif tag == "replace":
            rep += max(i2 - i1, j2 - j1)
        else:
            eq += i2 - i1
    return [
        inter / union, char_ratio, tok_ratio,
        float(ins), float(dele), float(rep), float(eq),
        float(len(ot)), float(len(nt)), float(len(nt) - len(ot)),
        float(abs(len(nt) - len(ot))),
        float(len(new) - len(old)), float(abs(len(new) - len(old))),
        float(len(sn - so)), float(len(so - sn)), float(inter),
        float(not ot), float(not nt),
        (len(nt) / len(ot)) if ot else 0.0,
    ]

# Appendix C, Table 8: the six transition-word groups; the feature is one count per group.
TRANSITION_GROUPS = {
    "reasoning": ("consequently", "clearly", "then", "furthermore", "additionally",
                  "moreover", "because", "besides", "also"),
    "evidence": ("illustration", "e.g.", "example", "instance", "specifically",
                 "demonstrate", "illustrate"),
    "rebuttal": ("however", "but", "yet", "although", "despite", "contrast",
                 "nevertheless", "nonetheless", "notwithstanding", "contrary",
                 "otherwise", "though"),
    "conclusion": ("therefore", "hence", "conclusion", "consideration", "indeed",
                   "finally", "lastly"),
    "details": ("specifically", "especially", "particular", "explain", "list",
                "enumerate", "detail", "namely", "including"),
    "causation": ("accordingly", "so", "because", "consequently", "hence", "since",
                  "therefore", "thus"),
}


def extract_v4():
    """The pinned construction (2026-08-11 subagent, L79): a unit is the set of rows sharing
    'Revision Index Level 0' (per the corpus toolkit's mergeUnit), purposes unioned across the
    unit, units with more than one distinct purpose DISCARDED (the paper's rule, not
    first-pick), many-to-many texts joined with spaces, no truncation."""
    from openpyxl import load_workbook                                # noqa: PLC0415

    units: dict = {}
    for wb_path in sorted(ANNOT.rglob("*.xlsx")):
        if wb_path.name.startswith("~$"):
            continue
        m = re.search(r"[\\/](12|23)[\\/]", str(wb_path))
        cyc = m.group(1) if m else "??"
        am = re.search(r"argrewrite_(\d+)", wb_path.stem)
        author = am.group(1) if am else wb_path.stem
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
        except Exception:
            continue

        def table(ws):
            ws.reset_dimensions()
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None
            hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

            def col(nm):
                return next((i for i, h in enumerate(hdr) if nm in h), None)

            ix = {"i": col("sentence index"), "txt": col("sentence content"),
                  "p": col("revision purpose level 0"),
                  "ri": col("revision index level 0")}
            if ix["i"] is None or ix["txt"] is None or ix["ri"] is None:
                return None
            return [{k: (r[v] if v is not None and v < len(r) else None)
                     for k, v in ix.items()} for r in rows[1:]]

        sheets = {ws.title.lower(): table(ws) for ws in wb.worksheets}
        wb.close()

        def as_i(v):
            try:
                return int(float(str(v).split(",")[0]))
            except (TypeError, ValueError):
                return None

        for kname, rows in sheets.items():
            if not rows:
                continue
            side = "old" if "old" in kname else "new"
            for r in rows:
                ri = as_i(r["ri"])
                if ri is None:
                    continue
                u = units.setdefault((str(wb_path), ri),
                                     {"author": author, "cycle": cyc, "purposes": [],
                                      "old": [], "new": [], "oi": [], "ni": []})
                txt = str(r["txt"] or "").strip()
                if txt:
                    u[side].append(txt)
                idx = as_i(r["i"])
                if idx is not None:
                    u["oi" if side == "old" else "ni"].append(idx)
                if r["p"]:
                    for p in str(r["p"]).lower().split(","):
                        p = p.strip()
                        if p and p not in ("none", "nan") and p not in u["purposes"]:
                            u["purposes"].append(p)

    events = []
    n_multi = 0
    for u in units.values():
        if not u["purposes"]:
            continue
        if len(u["purposes"]) > 1:
            n_multi += 1
            continue
        fine = FINE9.get(u["purposes"][0])
        if not fine:
            continue
        events.append({"author": u["author"], "cycle": u["cycle"],
                       "raw": u["purposes"][0], "fine": fine,
                       "binary": "surface" if fine in SURFACE9 else "content",
                       "old": " ".join(u["old"]), "new": " ".join(u["new"]),
                       "pos_idx": (min(u["ni"]) if u["ni"]
                                   else (min(u["oi"]) if u["oi"] else 0)),
                       "pos_old": min(u["oi"]) if u["oi"] else 0})
    from collections import Counter                                   # noqa: PLC0415
    cyc_n = Counter(e["cycle"] for e in events)
    cls_n = Counter(e["fine"] for e in events)
    print(f"v4 units: {len(events)} (paper 3,238); multi-purpose discarded {n_multi}; "
          f"per-cycle {dict(cyc_n)}")
    print(f"v4 per-class: {dict(cls_n)}")
    return events


def extract_raw():
    """Every (pair, purpose) event before label filtering or dedup — the hunt's raw stream."""
    from openpyxl import load_workbook                                # noqa: PLC0415

    events = []
    for wb_path in sorted(ANNOT.rglob("*.xlsx")):
        m = re.search(r"[\\/](12|23)[\\/]", str(wb_path))
        cyc = m.group(1) if m else "??"
        am = re.search(r"argrewrite_(\d+)", wb_path.stem)
        author = am.group(1) if am else wb_path.stem
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
        except Exception:
            continue

        def table(ws):
            ws.reset_dimensions()
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None
            hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

            def col(nm):
                return next((i for i, h in enumerate(hdr) if nm in h), None)

            ix = {"i": col("sentence index"), "txt": col("sentence content"),
                  "al": col("aligned index"), "p1": col("purpose level 1"),
                  "p0": col("purpose level 0")}
            if ix["i"] is None or ix["txt"] is None:
                return None
            out = []
            for r in rows[1:]:
                out.append({k: (r[v] if v is not None and v < len(r) else None)
                            for k, v in ix.items()})
            return out

        sheets = {ws.title.lower(): table(ws) for ws in wb.worksheets}
        wb.close()
        old = next((v for k, v in sheets.items() if "old" in k and v), None) or []
        new = next((v for k, v in sheets.items() if "new" in k and v), None) or []

        def as_int(v):
            try:
                return int(float(str(v).split(",")[0]))
            except (TypeError, ValueError):
                return None

        old_by_i = {}
        for r in old:
            i = as_int(r["i"])
            if i is not None:
                old_by_i[i] = r

        def purposes(*rows_):
            out = []
            for r in rows_:
                if not r:
                    continue
                raw = r["p1"] or r["p0"]
                if not raw:
                    continue
                for piece in str(raw).lower().split(","):
                    piece = piece.strip()
                    if piece and piece not in ("none", "nan"):
                        out.append(piece)
            seen, uniq = set(), []
            for p in out:
                if p not in seen:
                    seen.add(p)
                    uniq.append(p)
            return uniq

        consumed_old = set()
        for r in new:
            al = as_int(r["al"])
            o = old_by_i.get(al)
            if o is not None:
                consumed_old.add(al)
            for p in purposes(r, o):
                events.append({"author": author, "cycle": cyc, "raw": p,
                               "old": str((o or {}).get("txt") or "").strip()[:600],
                               "new": str(r["txt"] or "").strip()[:600],
                               "pos_idx": as_int(r["i"]) or 0})
        for i, r in old_by_i.items():
            if i in consumed_old:
                continue
            for p in purposes(r):
                events.append({"author": author, "cycle": cyc, "raw": p,
                               "old": str(r["txt"] or "").strip()[:600], "new": "",
                               "pos_idx": i})
    return events


def extract_v2():
    events = extract_raw()
    for e in events:
        e["fine"] = FINE9.get(e["raw"])
    events = [e for e in events if e["fine"]]
    # v3: one purpose per revision pair. v2's comma-splitting produced label-conflicting
    # duplicates of the same (old, new) pair, which is the prime suspect for the fine-task
    # collapse (0.27 vs the published 0.44+) -- the paper's n is 3,238 and the split gave 3,365.
    seen: set = set()
    first_only = []
    for e in events:
        key = (e["author"], e["cycle"], e["old"], e["new"])
        if key in seen:
            continue
        seen.add(key)
        first_only.append(e)
    events = first_only
    for e in events:
        e["binary"] = "surface" if e["fine"] in SURFACE9 else "content"
    from collections import Counter                                   # noqa: PLC0415
    cyc = Counter(e["cycle"] for e in events)
    print(f"v3 per-cycle n: {dict(cyc)} (total {len(events)}; paper sentential 3,238)")
    return events


def main() -> None:
    import argparse                                                   # noqa: PLC0415
    import numpy as np                                                # noqa: PLC0415
    from sklearn.model_selection import KFold                         # noqa: PLC0415
    from sklearn.metrics import f1_score, accuracy_score              # noqa: PLC0415
    from xgboost import XGBClassifier                                 # noqa: PLC0415

    ap = argparse.ArgumentParser()
    ap.add_argument("--drop-raw", default="",
                    help="comma-separated raw purpose labels to exclude (composition arms)")
    ap.add_argument("--tasks", default="binary,fine")
    ap.add_argument("--out", default="replication.json")
    ap.add_argument("--extract", default="v4", choices=["v4", "v2"],
                    help="v4 is the pinned construction (L79); v2 kept for provenance")
    ap.add_argument("--grid", action="store_true",
                    help="search the 36-point grid instead of the published footnote values")
    ap.add_argument("--balanced", action="store_true",
                    help="balanced sample weights in fit (the rare-class-gap suspect)")
    ap.add_argument("--diff-features", action="store_true",
                    help="augment USE with explicit change features: E_old-E_new, |diff|, "
                         "cosine (the concatenation-never-states-the-change suspect)")
    ap.add_argument("--change-features", action="store_true",
                    help="add the 19 surface string-diff features (L85): token Jaccard, "
                         "sequence-matcher ratios, insert/delete/replace counts, length "
                         "deltas, empty-side flags. Cheap, and they carry the binary task")
    ap.add_argument("--oversample", type=float, default=0.0,
                    help="pre-CV oversample factor for the five underrepresented classes; "
                         "1.49 tests the L79-supplement arithmetic (predicts majority .29, "
                         "word-usage F1 .45, rare classes .35 to .54)")
    args = ap.parse_args()

    events = extract_v4() if args.extract == "v4" else extract_v2()
    if args.drop_raw:
        drops = {s.strip().lower() for s in args.drop_raw.split(",") if s.strip()}
        before = len(events)
        events = [e for e in events if e["raw"] not in drops]
        print(f"drop-raw {sorted(drops)}: {before} -> {len(events)} examples")
    if args.oversample > 1.0:
        import random                                                 # noqa: PLC0415
        rng = random.Random(41)
        five = ("claim", "evidence", "organization", "precision", "rebuttal")
        extra = []
        for cls in five:
            pool = [e for e in events if e["fine"] == cls]
            k = round((args.oversample - 1.0) * len(pool))
            extra.extend(dict(e) for e in rng.choices(pool, k=k))
        events = events + extra
        print(f"oversampled the five small classes x{args.oversample}: "
              f"+{len(extra)} -> {len(events)} (duplicates enter the CV pool by design; "
              f"that leakage is the hypothesis under test)")
    print(f"extract {args.extract}: {len(events)} examples "
          f"(paper sentential n = 3,238), {len({e['author'] for e in events})} authors")

    # ── traditional features: length, position, POS tag TFs, transition TFs (old and new)
    import spacy                                                      # noqa: PLC0415
    nlp = spacy.load("en_core_web_sm", disable=["parser", "ner", "lemmatizer"])
    tagset = sorted({"ADJ", "ADP", "ADV", "AUX", "CCONJ", "DET", "INTJ", "NOUN", "NUM",
                     "PART", "PRON", "PROPN", "PUNCT", "SCONJ", "SYM", "VERB", "X"})

    def trad(e):
        if args.extract == "v4":
            # Their feature set: per sentence, length + position + POS-19 term frequency
            # + one count per transition group (six).
            row = []
            for txt, pos in ((e["old"], e.get("pos_old", 0)),
                             (e["new"], e.get("pos_idx", 0))):
                row.append(len(txt.split()))
                row.append(pos)
                doc = nlp(txt) if txt else []
                counts = {t: 0 for t in TAGSET19}
                for tok in doc:
                    if tok.pos_ in counts:
                        counts[tok.pos_] += 1
                row.extend(counts[t] for t in TAGSET19)
                low = " " + txt.lower() + " "
                for grp in TRANSITION_GROUPS.values():
                    row.append(sum(low.count(w) if w.endswith(".")
                                   else low.count(" " + w + " ") for w in grp))
            return row
        row = [len(e["old"].split()), len(e["new"].split()), e["pos_idx"]]
        for txt in (e["old"], e["new"]):
            doc = nlp(txt) if txt else []
            counts = {t: 0 for t in tagset}
            for tok in doc:
                if tok.pos_ in counts:
                    counts[tok.pos_] += 1
            row.extend(counts[t] for t in tagset)
            low = " " + txt.lower() + " "
            row.extend(low.count(" " + w + " ") for w in TRANSITIONS)
        return row

    print("featurizing (spaCy POS + transitions)...", flush=True)
    X_trad = np.array([trad(e) for e in events], float)
    X_change = None
    if args.change_features:
        X_change = np.array([change_features(e["old"], e["new"]) for e in events], float)
        X_trad = np.hstack([X_trad, X_change])
        print(f"change features on: X_trad now {X_trad.shape[1]}-dim", flush=True)

    # ── USE embeddings of the <old, new> pair, concatenated
    print("loading Universal Sentence Encoder (transformer variant)...", flush=True)
    import os                                                         # noqa: PLC0415
    os.environ.setdefault("TFHUB_CACHE_DIR", str(REPO / "results" / "use_cache"))
    os.environ["CUDA_VISIBLE_DEVICES"] = "-1"   # TF on CPU; the card belongs to torch and ollama
    import tensorflow_hub as hub                                      # noqa: PLC0415
    use = hub.load("https://tfhub.dev/google/universal-sentence-encoder-large/5")

    def embed(texts):
        out = []
        for i in range(0, len(texts), 256):
            out.append(use(texts[i:i + 256]).numpy())
        return np.vstack(out)

    print("embedding old/new sentences...", flush=True)
    E_old = embed([e["old"] or " " for e in events])
    E_new = embed([e["new"] or " " for e in events])
    if args.diff_features:
        d = E_old - E_new
        denom = np.linalg.norm(E_old, axis=1) * np.linalg.norm(E_new, axis=1)
        cos = (np.sum(E_old * E_new, axis=1) / np.where(denom == 0, 1, denom))[:, None]
        X_use = np.hstack([E_old, E_new, d, np.abs(d), cos])
    else:
        X_use = np.hstack([E_old, E_new])

    GRID = [{"n_estimators": n, "max_depth": d, "learning_rate": lr}
            for n in (250, 500, 750, 1000) for d in (3, 4, 5) for lr in (.1, .05, .01)]

    FIXED = {"binary": {"n_estimators": 500, "max_depth": 4, "learning_rate": .05},
             "fine": {"n_estimators": 750, "max_depth": 5, "learning_rate": .05}}

    def evaluate(task):
        from sklearn.utils.class_weight import compute_sample_weight  # noqa: PLC0415
        y_raw = [e[task] for e in events]
        classes = sorted(set(y_raw))
        y = np.array([classes.index(v) for v in y_raw])
        arms = {"majority": None, "features": X_trad, "use": X_use,
                "features_use": np.hstack([X_trad, X_use])}
        if X_change is not None:
            arms["change_only"] = X_change
        res = {}
        kf = KFold(n_splits=5, shuffle=True, random_state=42)
        for arm, X in arms.items():
            if arm == "majority":
                f1s, accs = [], []
                for tr, te in kf.split(np.zeros(len(y))):
                    maj = np.bincount(y[tr]).argmax()
                    preds = np.full_like(y[te], maj)
                    f1s.append(f1_score(y[te], preds, average="macro"))
                    accs.append(accuracy_score(y[te], preds))
                res[arm] = {"f1": float(np.mean(f1s)), "acc": float(np.mean(accs))}
                print(f"  {task}/majority: F1 {res[arm]['f1']:.3f} acc "
                      f"{res[arm]['acc']:.3f} (target {TARGETS[task][arm]})", flush=True)
                continue
            search = GRID if args.grid else [FIXED[task]]
            best = None
            oof = np.empty(len(y), dtype=int) if len(search) == 1 else None
            for g in search:
                f1s, accs = [], []
                for tr, te in kf.split(X):
                    # n_jobs fixed, not -1: XGBoost is thread-order dependent above one
                    # thread, so -1 tied every number to whatever else the host was running.
                    clf = XGBClassifier(**g, tree_method="hist", n_jobs=8,
                                        eval_metric="mlogloss", verbosity=0)
                    sw = (compute_sample_weight("balanced", y[tr])
                          if args.balanced else None)
                    clf.fit(X[tr], y[tr], sample_weight=sw)
                    p = clf.predict(X[te])
                    if oof is not None:
                        oof[te] = p
                    f1s.append(f1_score(y[te], p, average="macro"))
                    accs.append(accuracy_score(y[te], p))
                cand = {"f1": float(np.mean(f1s)), "acc": float(np.mean(accs)), "grid": g}
                if best is None or cand["f1"] > best["f1"]:
                    best = cand
            if oof is not None:
                pc = f1_score(y, oof, average=None,
                              labels=list(range(len(classes))), zero_division=0)
                best["per_class_f1"] = {classes[i]: round(float(v), 3)
                                        for i, v in enumerate(pc)}
            res[arm] = best
            print(f"  {task}/{arm}: F1 {best['f1']:.3f} acc {best['acc']:.3f} "
                  f"(target {TARGETS[task][arm]})", flush=True)
        return res

    run_tasks = [t.strip() for t in args.tasks.split(",") if t.strip() in TARGETS]
    out = {"n": len(events), "drop_raw": args.drop_raw, "tasks": run_tasks,
           "extract": args.extract, "grid": bool(args.grid),
           "balanced": bool(args.balanced), "diff_features": bool(args.diff_features),
           "oversample": args.oversample, "targets": TARGETS, "results": {},
           "fine_majority_note": (
               "the printed fine Majority row (.05/.29) contradicts the paper's own "
               "Table 4 (word-usage 1,030 of 3,238 implies .05/.32); we gate the "
               "majority arm against the table-implied values (L79)")}
    for task in run_tasks:
        print(f"== {task}", flush=True)
        out["results"][task] = evaluate(task)

    deltas, passed = {}, True
    for task in run_tasks:
        for arm, (tf1, tacc) in TARGETS[task].items():
            r = out["results"][task][arm]
            deltas[f"{task}/{arm}"] = {"f1": round(r["f1"] - tf1, 3),
                                       "acc": round(r["acc"] - tacc, 3)}
            if arm == "features_use" and (abs(round(r["f1"], 2) - tf1) > 0.005 or
                                          abs(round(r["acc"], 2) - tacc) > 0.005):
                passed = False
    out["deltas"] = deltas
    out["verdict"] = "CAPABILITY-PASS" if passed else "NOT-MATCHED"
    print(f"\n  >>> {out['verdict']}\n  deltas: {json.dumps(deltas, indent=1)}")

    RESULTS.mkdir(parents=True, exist_ok=True)
    (RESULTS / args.out).write_text(json.dumps(out, indent=1),
                                    encoding="utf-8", newline="\n")
    print(f"wrote {(RESULTS / args.out).relative_to(REPO)}")


if __name__ == "__main__":
    main()
